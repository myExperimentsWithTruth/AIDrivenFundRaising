"""Build a dummy Fundraise_Reconciliation_Sheet.xlsx that mirrors the production layout.
All data is synthetic. Names are obviously generic; UTRs and amounts are made up.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/sessions/affectionate-modest-heisenberg/mnt/KD Fundraise/code/examples/Fundraise_Reconciliation_Sheet.xlsx"

# ---------- styling helpers ----------
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="44546A")
THIN = Side(border_style="thin", color="D6D3D1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BODY_FONT = Font(name="Arial", size=10)
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="78716C")
DASH_LABEL = Font(name="Arial", size=11, bold=True)
ACCENT_FILL = PatternFill("solid", start_color="FEF3C7")  # amber-100

def style_header(sheet, cols):
    for c in range(1, cols + 1):
        cell = sheet.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

def style_body(sheet, rows, cols, start_row=2):
    for r in range(start_row, start_row + rows):
        for c in range(1, cols + 1):
            cell = sheet.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER

def set_widths(sheet, widths):
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w

# ---------- workbook ----------
wb = Workbook()
wb.remove(wb.active)

# ---------- Master_List ----------
ml = wb.create_sheet("Master_List")
ml_cols = ["#", "Name", "Batch / Group", "Pledged_INR", "Received_Confirmed",
           "Received_Pending", "Total_Claimed", "Payment Status",
           "Verification Status", "Bank_Refs", "Pending_Refs", "Notes"]
ml.append(ml_cols)

# 10 fake contributors
donors = [
    (1, "Anand S",   "Batch-1998",  50000, "OK",      "Verified",             "UPI/2026051912345",                            "",                             ""),
    (2, "Bhavna K",  "Work-circle", 25000, "OK",      "Verified",             "UPI/2026051923456",                            "",                             ""),
    (3, "Chetan M",  "Batch-1998", 100000, "OK",      "Verified",             "UPI/2026051967890",                            "",                             ""),
    (4, "Deepa R",   "Family",      10000, "PENDING", "Pending Verification", "",                                             "Screenshot · WhatsApp · 21-May 14:20", "Claim sent, not yet in bank"),
    (5, "Eshan P",   "Batch-2000",  15000, "PARTIAL", "Verified",             "UPI/2026052156789",                            "Screenshot · WhatsApp · 22-May 09:10", "10K in bank, 5K pending"),
    (6, "Farhan A",  "Overseas",    20000, "OK",      "Verified",             "MLP/MILAAP-EXT-9921",                          "",                             "Routed via Milaap to Holding"),
    (7, "Gaurav S",  "Batch-1998",   5000, "OK",      "Verified",             "HOLD/H001",                                    "",                             "Paid into Holding intermediary"),
    (8, "Hema V",    "Work-circle", 30000, "PENDING", "Pending Verification", "",                                             "Screenshot · WhatsApp · 21-May 19:42", "Awaiting next PNB statement"),
    (9, "Ishan T",   "Batch-1999",  50000, "OK",      "Verified",             "UPI/2026052045678",                            "",                             ""),
    (10, "Jaya M",   "Family",      25000, "OK",      "Verified",             "UPI/2026052078901",                            "",                             ""),
]

for d in donors:
    n, name, batch, pledged, status, verif, bank_refs, pend_refs, notes = d
    ml.append([
        n, name, batch, pledged,
        # E: Received_Confirmed via SUMIFS across all bank sheets
        f'=SUMIFS(PNB_Statement!C:C,PNB_Statement!G:G,A{n+1},PNB_Statement!D:D,"CR")'
        f'+SUMIFS(Canara_Statement!C:C,Canara_Statement!G:G,A{n+1},Canara_Statement!D:D,"CR")'
        f'+SUMIFS(Holding_Account!C:C,Holding_Account!G:G,A{n+1},Holding_Account!D:D,"CR")',
        # F: Received_Pending — manual, hardcoded here for dummy
        {4: 10000, 5: 5000, 8: 30000}.get(n, 0),
        # G: Total_Claimed
        f"=E{n+1}+F{n+1}",
        status, verif, bank_refs, pend_refs, notes
    ])

set_widths(ml, [5, 18, 14, 14, 18, 16, 14, 14, 20, 30, 36, 30])
style_header(ml, len(ml_cols))
style_body(ml, len(donors), len(ml_cols))
ml.freeze_panes = "A2"

# Currency formatting on numeric cols
for r in range(2, 2 + len(donors)):
    for col in ("D", "E", "F", "G"):
        ml[f"{col}{r}"].number_format = '#,##0;(#,##0);"-"'

# ---------- Bank sheet helper ----------
BANK_COLS = ["Date", "Time", "Amount", "Type", "Sender (bank-name)", "UTR / UPI Ref",
             "Mapped_To_#", "Status", "Source", "Linked Name", "Notes"]

def build_bank_sheet(name, rows):
    s = wb.create_sheet(name)
    s.append(BANK_COLS)
    for r in rows:
        s.append(list(r))
    # Linked Name formula (col J) — VLOOKUP on Mapped_To_# (col G)
    for r in range(2, 2 + len(rows)):
        s[f"J{r}"] = f'=IFERROR(VLOOKUP(G{r},Master_List!A:B,2,FALSE),"")'
    set_widths(s, [12, 8, 12, 8, 30, 28, 12, 14, 10, 18, 30])
    style_header(s, len(BANK_COLS))
    style_body(s, len(rows), len(BANK_COLS))
    s.freeze_panes = "A2"
    for r in range(2, 2 + len(rows)):
        s[f"C{r}"].number_format = '#,##0;(#,##0);"-"'
    return s

# ---------- PNB_Statement (date-only, no time field) ----------
pnb_rows = [
    ("19-05-2026", "", 50000, "CR", "ANAND SHARMA",       "UPI/2026051912345", 1, "Matched",    "Stmt", "", ""),
    ("19-05-2026", "", 25000, "CR", "BHAVNA KAPOOR",      "UPI/2026051923456", 2, "Matched",    "Stmt", "", ""),
    ("20-05-2026", "", 15000, "CR", "RAVI KUMAR",         "UPI/2026052034567", "", "Unreviewed","Stmt", "", "Sender not on Master_List — ask fundraiser"),
    ("20-05-2026", "", 50000, "CR", "ISHAN TANDON",       "UPI/2026052045678", 9, "Matched",    "Stmt", "", ""),
    ("21-05-2026", "", 10000, "CR", "ESHAN PAUL",         "UPI/2026052156789", 5, "Matched",    "Stmt", "", "First instalment; 5K still pending"),
]
build_bank_sheet("PNB_Statement", pnb_rows)

# ---------- Canara_Statement (with time) ----------
canara_rows = [
    ("19-05-2026", "14:32", 100000, "CR", "CHETAN MEHTA",        "UPI/2026051967890", 3, "Matched",  "Stmt", "", ""),
    ("20-05-2026", "10:15",  25000, "CR", "JAYA MENON",          "UPI/2026052078901", 10, "Matched", "Stmt", "", ""),
    ("20-05-2026", "16:42",  25000, "CR", "HOLDING POOL SWEEP",  "SWP/HOLD-CAN-001",  "", "Internal","Stmt", "", "Sweep from Holding_Account row 4 — NOT a contribution"),
]
build_bank_sheet("Canara_Statement", canara_rows)

# ---------- Holding_Account (intermediary pool) ----------
hold_rows = [
    ("18-05-2026", "11:20",  5000, "CR", "GAURAV SINGH",          "HOLD/H001",        7, "Matched",  "Stmt", "", ""),
    ("18-05-2026", "13:45", 20000, "CR", "FARHAN AHMED (MILAAP)", "MLP/MILAAP-EXT-9921", 6, "Matched","Stmt", "", "Routed via Milaap into Holding"),
    ("20-05-2026", "16:42", 25000, "DR", "SWEEP TO CANARA",       "SWP/HOLD-CAN-001", "", "Internal", "Stmt", "", "SWEEP — Holding → Canara. NOT a contribution."),
]
build_bank_sheet("Holding_Account", hold_rows)

# Add Holding balance cell (header section above data would be ideal, but for simplicity put at the side)
hold = wb["Holding_Account"]
hold["M1"] = "Holding Balance"
hold["M1"].font = DASH_LABEL
hold["N1"] = '=SUMIFS(C:C,D:D,"CR")-SUMIFS(C:C,D:D,"DR")'
hold["N1"].font = Font(name="Arial", size=11, bold=True)
hold["N1"].number_format = '#,##0;(#,##0);"-"'
hold.column_dimensions["M"].width = 18
hold.column_dimensions["N"].width = 14

# ---------- Distribution_List ----------
dl = wb.create_sheet("Distribution_List")
dl_cols = ["#", "Display Name", "Phone", "Tier", "Master_List Link",
           "Notes", "Last Sent Date", "Last Send Status", "Removed"]
dl.append(dl_cols)
recipients = [
    (1,  "Anand S",        "",            "Existing",   1,  "Saved contact",                              "2026-05-22", "msg-dblcheck", "FALSE"),
    (2,  "Bhavna K",       "",            "Existing",   2,  "Saved contact",                              "2026-05-22", "msg-dblcheck", "FALSE"),
    (3,  "Chetan M",       "",            "Existing",   3,  "",                                           "2026-05-22", "msg-check",    "FALSE"),
    (4,  "Deepa R",        "918XXXXX1234","Existing",   4,  "Unsaved contact — using deep-link",          "2026-05-22", "msg-dblcheck", "FALSE"),
    (5,  "Eshan P",        "",            "Existing",   5,  "",                                           "2026-05-22", "msg-check",    "FALSE"),
    (6,  "Farhan A",       "971XXXXX5678","Existing",   6,  "Overseas — Milaap contributor",              "2026-05-22", "msg-dblcheck", "FALSE"),
    (7,  "Gaurav S",       "",            "Existing",   7,  "Via Holding intermediary",                   "2026-05-22", "msg-check",    "FALSE"),
    (8,  "Hema V",         "",            "Existing",   8,  "",                                           "2026-05-22", "msg-dblcheck", "FALSE"),
    (9,  "Ishan T",        "",            "Existing",   9,  "",                                           "2026-05-22", "msg-check",    "FALSE"),
    (10, "Jaya M",         "",            "Existing",   10, "",                                           "2026-05-22", "msg-dblcheck", "FALSE"),
    (11, "Karthik N",      "919XXXXX7777","Added 21-May","",  "Forwarder — wants future updates",         "2026-05-22", "msg-check",    "FALSE"),
    (12, "Lakshmi V",      "",            "Existing",   "",  "Forwarder; circulates within family group", "2026-05-22", "msg-dblcheck", "FALSE"),
    (13, "Manish K",       "",            "Existing",   "",  "Not a contributor; forwarder only",         "2026-05-22", "msg-check",    "FALSE"),
    (14, "Batch-1998 group","",           "Group",      "",  "WhatsApp group — manual touch needed",      "2026-05-22", "msg-check",    "FALSE"),
    (15, "Nikhil P",       "",            "Existing",   "",  "Asked to be removed on 21-May",             "2026-05-21", "msg-dblcheck", "TRUE"),
]
for r in recipients:
    dl.append(list(r))

set_widths(dl, [5, 22, 16, 16, 14, 38, 16, 16, 10])
style_header(dl, len(dl_cols))
style_body(dl, len(recipients), len(dl_cols))
dl.freeze_panes = "A2"

# ---------- Audit_Log ----------
al = wb.create_sheet("Audit_Log")
al_cols = ["Timestamp", "Master_List #", "Sheet", "Row", "Field",
           "Before", "After", "Trigger", "Evidence Ref"]
al.append(al_cols)
audit_entries = [
    ("2026-05-19T08:14:22", 1,  "PNB_Statement",   2, "Mapped_To_#",         "",     "1",     "Bank stmt import",    "UPI/2026051912345"),
    ("2026-05-19T08:14:35", 2,  "PNB_Statement",   3, "Mapped_To_#",         "",     "2",     "Bank stmt import",    "UPI/2026051923456"),
    ("2026-05-19T08:14:48", 3,  "Canara_Statement",2, "Mapped_To_#",         "",     "3",     "Bank stmt import",    "UPI/2026051967890"),
    ("2026-05-20T09:02:11", 4,  "Master_List",     5, "Received_Pending",    "0",    "10000", "Donor screenshot",    "WhatsApp · 21-May 14:20"),
    ("2026-05-20T15:42:55", "", "Canara_Statement",4, "Status",              "",     "Internal","Sweep handler",     "SWP/HOLD-CAN-001"),
    ("2026-05-21T07:55:12", "", "daily_updates",   0, "Generated",           "",     "Verified=200000,Pend=45000",   "Generate §8.2", "daily_updates/20260521_update.md"),
    ("2026-05-21T07:57:33", "", "audit",           0, "Verification PASS",   "",     "",     "Verification §8.1",    "headline matches Excel"),
    ("2026-05-21T07:58:02", "", "audit",           0, "Audit APPROVED",      "",     "",     "Audit agent §9.2",     "archive/audit_20260521_075800.md"),
    ("2026-05-22T03:08:19", 11, "Distribution_List",12,"Added recipient",    "",     "Karthik N", "Add new recipient", "Forwarder request"),
    ("2026-05-22T03:21:44", 15, "Distribution_List",16,"Removed",             "FALSE","TRUE", "Recipient opt-out",   "WhatsApp 21-May 22:30"),
]
for e in audit_entries:
    al.append(list(e))

set_widths(al, [22, 14, 18, 8, 24, 14, 22, 22, 38])
style_header(al, len(al_cols))
style_body(al, len(audit_entries), len(al_cols))
al.freeze_panes = "A2"

# ---------- Action_Items ----------
ai = wb.create_sheet("Action_Items")
ai_cols = ["#", "Created", "Type", "Pointer", "Summary", "Status", "Resolved by"]
ai.append(ai_cols)
actions = [
    (1, "2026-05-20T09:35:18", "Unreviewed",   "PNB_Statement!row=4", "RAVI KUMAR · ₹15,000 · 20-May · do we recognise this sender?",       "Open",     ""),
    (2, "2026-05-21T07:30:09", "Stale-Pending","Master_List!#=8",     "Hema V · ₹30,000 pending > 48 hours · investigate",                  "Open",     ""),
    (3, "2026-05-22T03:08:19", "New-Recipient","Distribution_List!#=11","Forwarder Karthik N added; resend today's update once verified.","Resolved", "Distribution_List row 12"),
]
for a in actions:
    ai.append(list(a))

set_widths(ai, [5, 22, 16, 22, 60, 12, 22])
style_header(ai, len(ai_cols))
style_body(ai, len(actions), len(ai_cols))
ai.freeze_panes = "A2"

# ---------- Dashboard ----------
db = wb.create_sheet("Dashboard")
db["A1"] = "Fundraise Dashboard"
db["A1"].font = Font(name="Arial", size=14, bold=True, color="92400E")
db.merge_cells("A1:D1")

db["A3"] = "Snapshot as of"
db["A3"].font = DASH_LABEL
db["B3"] = "22-May-2026 03:30 IST"
db["B3"].font = BODY_FONT

# Headline three-bucket totals
db["A5"] = "Headline (three buckets)"
db["A5"].font = DASH_LABEL

db["A6"] = "Verified in bank (named)"
db["B6"] = "=SUM(Master_List!E:E)"
db["A7"] = "Awaiting bank statement (donor evidence in)"
db["B7"] = "=SUM(Master_List!F:F)"
db["A8"] = "Awaiting identification (bank credits without name)"
db["B8"] = ('=SUMIFS(PNB_Statement!C:C,PNB_Statement!G:G,"",PNB_Statement!D:D,"CR",PNB_Statement!H:H,"Unreviewed")'
            '+SUMIFS(Canara_Statement!C:C,Canara_Statement!G:G,"",Canara_Statement!D:D,"CR",Canara_Statement!H:H,"Unreviewed")')

db["A9"] = "Grand total (publishable headline)"
db["A9"].font = DASH_LABEL
db["B9"] = "=B6+B7+B8"
db["B9"].font = Font(name="Arial", size=11, bold=True)
db["A9"].fill = ACCENT_FILL
db["B9"].fill = ACCENT_FILL

for r in (6, 7, 8, 9):
    db[f"B{r}"].number_format = '"₹"#,##0;("₹"#,##0);"-"'

# Channel breakdown
db["A11"] = "By channel (gross CR)"
db["A11"].font = DASH_LABEL
db["A12"] = "PNB_Statement"
db["B12"] = '=SUMIFS(PNB_Statement!C:C,PNB_Statement!D:D,"CR")'
db["A13"] = "Canara_Statement"
db["B13"] = '=SUMIFS(Canara_Statement!C:C,Canara_Statement!D:D,"CR")'
db["A14"] = "Holding_Account"
db["B14"] = '=SUMIFS(Holding_Account!C:C,Holding_Account!D:D,"CR")'

for r in (12, 13, 14):
    db[f"B{r}"].number_format = '"₹"#,##0;("₹"#,##0);"-"'

# Holding balance
db["A16"] = "Holding balance (residual in pool)"
db["A16"].font = DASH_LABEL
db["B16"] = '=SUMIFS(Holding_Account!C:C,Holding_Account!D:D,"CR")-SUMIFS(Holding_Account!C:C,Holding_Account!D:D,"DR")'
db["B16"].number_format = '"₹"#,##0;("₹"#,##0);"-"'

# Contributor status counts
db["A18"] = "Contributors by Payment Status"
db["A18"].font = DASH_LABEL
statuses = [("OK", 19), ("PARTIAL", 20), ("PENDING", 21), ("OVER", 22), ("REMOVED", 23)]
for status, r in statuses:
    db[f"A{r}"] = status
    db[f"B{r}"] = f'=COUNTIF(Master_List!H:H,"{status}")'

db["A25"] = "Open Action_Items"
db["A25"].font = DASH_LABEL
db["B25"] = '=COUNTIF(Action_Items!F:F,"Open")'

set_widths(db, [42, 22, 20, 20])

# Borders / fonts for dashboard rows
for r in range(5, 26):
    db[f"A{r}"].border = BORDER
    db[f"B{r}"].border = BORDER

# Reorder sheets
order = ["Master_List", "PNB_Statement", "Canara_Statement", "Holding_Account",
         "Distribution_List", "Audit_Log", "Action_Items", "Dashboard"]
wb._sheets = [wb[name] for name in order]

wb.save(OUT)
print(f"Saved {OUT}")
